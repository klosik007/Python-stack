import openpyxl, sys
from openpyxl.styles import Font

if len(sys.argv) < 2:
    print('Usage: xlsopen.py location')
    sys.exit()

N = sys.argv[1]
M = sys.argv[2]
file = sys.argv[3]

#print(N + '\n' + M + '\n' + file)

wb = openpyxl.load_workbook(file)
sheet = wb.get_active_sheet()

wb2 = openpyxl.Workbook()
sheet2 = wb2.get_active_sheet()

for row in range(1, int(N)):
    for col in range(1, sheet.max_column+1):
        sheet2.cell(row=row, column=col).value = sheet.cell(row=row, column=col).value

for row in range(int(N)+int(M), sheet.max_row+int(M)+1):
    for col in range(1, sheet.max_column+1):
        sheet2.cell(row=row, column=col).value = sheet.cell(row=row-int(M), column=col).value

wb2.save(('new_blank_r_ins.xlsx'))
