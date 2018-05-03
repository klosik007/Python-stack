import openpyxl, sys
#from openpyxl import get_column_letter, column_index_from_string
from openpyxl.styles import Font, NamedStyle
from openpyxl.chart import Reference, Series, BarChart

#wb = openpyxl.load_workbook('produceSales.xlsx')
#print(type(wb))
#print(wb.get_sheet_names())
#sheet = wb.get_sheet_by_name('Sheet')
'''print(sheet['A1'].value)
c = sheet['B1']
print(c.coordinate)
print(sheet.cell(row=2, column=4).value)
#print(sheet.max_row)'''

'''print(tuple(sheet['A1':'C3']))
for rowOfCellObjects in sheet['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('---END OF ROW---')'''

'''PRICE_UPDATES = {'Garlic': 3.02,
                 'Celery' : 1.20,
                 'Lemon' : 1.27}

for rowNum in range(2, sheet.max_row):
    produceName = sheet.cell(row = rowNum, column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column =2).value = PRICE_UPDATES[produceName]

maxRow = sheet.max_row
sheet['E1'] = '=SUM(D2:D%d)' % maxRow

sheet.freeze_panes = 'A2'



wb.save('updatedProduceSales.xlsx')'''

'''wb = openpyxl.Workbook()
sheet = wb.get_active_sheet()
for i in range(1,11):
    sheet['A' + str(i)] = i

refObj = Reference(sheet, (1, 1), (10, 1))
seriesObj = Series(refObj, title='First series')

chartObj = BarChart()
chartObj.append(seriesObj)
chartObj.drawing.top = 50
chartObj.drawing.left = 100

chartObj.drawing.width = 300
chartObj.drawing.height = 200

sheet.add_chart(chartObj)
wb.save('sampleChart.xlsx')'''

if len(sys.argv) < 2:
    print('Usage: xlsopen.py location')
    sys.exit()
number = ' '.join(sys.argv[1:])

wb = openpyxl.Workbook()
sheet = wb.get_active_sheet()

font = Font(bold=True)

for rowNum in range(1, int(number)+1):
    for colNum in range(1, int(number)+1):
        sheet.cell(row=rowNum + 1, column=colNum + 1).value = rowNum * colNum
        sheet.cell(row=rowNum+1, column=1).font = font
        sheet.cell(row=rowNum+1, column=1).value = rowNum
        sheet.cell(row=1, column=colNum + 1).font = font
        sheet.cell(row=1, column=colNum+1).value = colNum


wb.save('multiplication.xlsx')





