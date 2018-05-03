import openpyxl, sys
from openpyxl.styles import Font

if len(sys.argv) < 2:
    print('Usage: xlsopen.py location')
    sys.exit()
number = ' '.join(sys.argv[1:])

wb = openpyxl.Workbook()
sheet = wb.get_active_sheet()

font = Font(bold=True)

for rowNum in range(1, int(number)+1):
    for colNum in range(1, int(number)+1):
        sheet.cell(row=rowNum + 1, column=colNum + 1).value = rowNum * colNum
        sheet.cell(row=rowNum+1, column=1).font = font
        sheet.cell(row=rowNum+1, column=1).value = rowNum
        sheet.cell(row=1, column=colNum + 1).font = font
        sheet.cell(row=1, column=colNum+1).value = colNum

wb.save('multiplication.xlsx')