import openpyxl



wb = openpyxl.load_workbook('multiplication.xlsx')
sheet = wb.get_active_sheet()

wb2 = openpyxl.Workbook()
sheet2 = wb2.get_active_sheet()

sh2=[[0 for column in range(sheet.max_column+1)] for row in range(sheet.max_row+1)]

for row in range(1, sheet.max_row+1):
    for col in range(1, sheet.max_column+1):
        #sheet2.cell(row=col, column=row).value = sheet.cell(row=row, column=col)
        sh2[row][col] = sheet.cell(row=row, column=col).value

#print(sh2[1][6])

for row in range(1, sheet.max_column+1):
    for col in range(1, sheet.max_row + 1):
        sheet2.cell(row=row, column=col).value = sh2[col][row]

wb2.save('multi2.xlsx')